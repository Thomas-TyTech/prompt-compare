#!/usr/bin/env python3
"""
Convert Multi-Prompt Evaluation Results to Excel
Creates an Excel file with side-by-side comparison of prompt responses.
"""

import json
import pandas as pd
import argparse
import os
from datetime import datetime
from typing import Dict, Any, List


def convert_multi_prompt_to_excel(results_file: str, output_file: str = None) -> str:
    """
    Convert multi-prompt evaluation results to Excel with side-by-side comparison.
    """
    
    print(f"📊 Loading multi-prompt evaluation results from {results_file}...")
    
    try:
        with open(results_file, 'r', encoding='utf-8') as f:
            data = json.load(f)
    except Exception as e:
        print(f"❌ Error loading results file: {e}")
        return None
    
    session = data.get('evaluation_session', {})
    detailed_results = data.get('detailed_results', {})
    summary = data.get('summary', {})
    
    # Extract questions and responses
    questions_data = {}
    prompt_names = {}
    
    for prompt_key, prompt_data in detailed_results.items():
        if isinstance(prompt_data, dict) and 'detailed_results' in prompt_data:
            prompt_info = prompt_data.get('prompt_version', {})
            prompt_name = prompt_info.get('name', prompt_key)
            prompt_names[prompt_key] = prompt_name
            
            for result in prompt_data['detailed_results']:
                question_id = result.get('question_id', 'Unknown')
                question_text = result.get('question', 'No question text')
                response = result.get('response', 'No response')
                response_time = result.get('response_time_ms', 0)
                links_found = result.get('links_found', 0)
                links_valid = result.get('links_valid', 0)
                links_invalid = result.get('links_invalid', 0)
                category = result.get('category', 'general')
                complexity = result.get('complexity', 'basic')
                
                # Extract links
                link_validation_results = result.get('link_validation_results', [])
                valid_links = [link['url'] for link in link_validation_results if link.get('status') == 'valid']
                invalid_links = [link['url'] for link in link_validation_results if link.get('status') == 'invalid']
                warning_links = [link['url'] for link in link_validation_results if link.get('status') == 'warning']
                
                if question_id not in questions_data:
                    questions_data[question_id] = {
                        'question': question_text,
                        'category': category,
                        'complexity': complexity,
                        'responses': {}
                    }
                
                questions_data[question_id]['responses'][prompt_key] = {
                    'response': response,
                    'response_time_ms': response_time,
                    'links_found': links_found,
                    'links_valid': links_valid,
                    'links_invalid': links_invalid,
                    'valid_links': valid_links,
                    'invalid_links': invalid_links,
                    'warning_links': warning_links,
                    'prompt_name': prompt_name
                }
    
    # Sort questions by ID
    sorted_questions = sorted(questions_data.items(), key=lambda x: x[0])
    
    print(f"✅ Processing {len(sorted_questions)} questions across {len(prompt_names)} prompts...")
    
    # Prepare data for Excel - Clean side-by-side comparison
    comparison_data = []
    
    # Get prompt names in order (assuming 2 prompts: A and B)
    prompt_keys = list(prompt_names.keys())
    prompt_a_key = prompt_keys[0] if len(prompt_keys) > 0 else None
    prompt_b_key = prompt_keys[1] if len(prompt_keys) > 1 else None
    
    for question_id, question_data in sorted_questions:
        # Get response data for both prompts
        prompt_a_data = question_data['responses'].get(prompt_a_key, {}) if prompt_a_key else {}
        prompt_b_data = question_data['responses'].get(prompt_b_key, {}) if prompt_b_key else {}
        
        # Extract links for each prompt
        def format_links(response_data):
            valid_links = response_data.get('valid_links', [])
            invalid_links = response_data.get('invalid_links', [])
            warning_links = response_data.get('warning_links', [])
            
            all_links = []
            if valid_links:
                all_links.extend(valid_links)
            if warning_links:
                all_links.extend(warning_links)
            if invalid_links:
                all_links.extend(invalid_links)
            
            return "\n".join(all_links) if all_links else ""
        
        row = {
            'Question': question_data['question'],
            'Answer A': prompt_a_data.get('response', ''),
            'Answer B': prompt_b_data.get('response', ''),
            'Links A': format_links(prompt_a_data),
            'Links B': format_links(prompt_b_data)
        }
        
        comparison_data.append(row)
    
    # Create DataFrame
    df_comparison = pd.DataFrame(comparison_data)
    
    # Prepare summary data
    summary_data = []
    if summary and 'prompt_comparison' in summary:
        for prompt_name, metrics in summary['prompt_comparison'].items():
            summary_data.append({
                'Prompt_Version': prompt_name,
                'Total_Links': metrics.get('total_links', 0),
                'Valid_Links': metrics.get('valid_links', 0),
                'Invalid_Links': metrics.get('invalid_links', 0),
                'Link_Success_Rate_%': metrics.get('link_success_rate', 0),
                'Avg_Response_Time_ms': metrics.get('avg_response_time_ms', 0),
                'Successful_API_Calls': metrics.get('successful_api_calls', 0),
                'Failed_API_Calls': metrics.get('failed_api_calls', 0),
                'Questions_with_Invalid_Links': metrics.get('questions_with_invalid_links', 0)
            })
    
    df_summary = pd.DataFrame(summary_data)
    
    # Generate output filename if not provided
    if not output_file:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        base_name = session.get('name', 'multi_prompt_evaluation').replace(' ', '_')
        output_file = f"{base_name}_{timestamp}.xlsx"
    
    try:
        # Create Excel writer with options for better formatting
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Write comparison results
            df_comparison.to_excel(writer, sheet_name='Question_Comparison', index=False)
            
            # Write summary
            if not df_summary.empty:
                df_summary.to_excel(writer, sheet_name='Summary', index=False)
            
            # Write session metadata
            metadata_data = [
                ['Session Name', session.get('name', 'N/A')],
                ['Session Description', session.get('description', 'N/A')],
                ['Session ID', session.get('id', 'N/A')],
                ['Created At', session.get('created_at', 'N/A')],
                ['Total Questions', len(sorted_questions)],
                ['Prompt Versions', len(prompt_names)],
                ['Generated On', datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
            ]
            
            df_metadata = pd.DataFrame(metadata_data, columns=['Field', 'Value'])
            df_metadata.to_excel(writer, sheet_name='Metadata', index=False)
            
            # Get workbooks for formatting
            workbook = writer.book
            
            # Format comparison sheet for clean, uniform appearance
            if 'Question_Comparison' in writer.sheets:
                worksheet = writer.sheets['Question_Comparison']
                
                from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
                
                # Set uniform column widths for clean rectangular appearance
                column_widths = {
                    'A': 60,  # Question
                    'B': 80,  # Answer A  
                    'C': 80,  # Answer B
                    'D': 60,  # Links A
                    'E': 60   # Links B
                }
                
                for col_letter, width in column_widths.items():
                    worksheet.column_dimensions[col_letter].width = width
                
                # Set uniform row height for clean appearance
                for row in range(1, len(comparison_data) + 2):  # +2 for header
                    worksheet.row_dimensions[row].height = 120 if row > 1 else 25  # Taller rows for content, normal for header
                
                # Create border style
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'), 
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # Header formatting
                header_fill = PatternFill(start_color='E6F3FF', end_color='E6F3FF', fill_type='solid')
                header_font = Font(bold=True, size=12)
                
                # Apply formatting to all cells
                for row in worksheet.iter_rows():
                    for cell in row:
                        # Apply borders to all cells
                        cell.border = thin_border
                        
                        # Header row formatting
                        if cell.row == 1:
                            cell.fill = header_fill
                            cell.font = header_font
                            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        else:
                            # Content formatting
                            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                            
                            # Different background for alternating columns
                            if cell.column in [2, 4]:  # Answer A and Links A columns
                                cell.fill = PatternFill(start_color='F8FFFE', end_color='F8FFFE', fill_type='solid')
                            elif cell.column in [3, 5]:  # Answer B and Links B columns  
                                cell.fill = PatternFill(start_color='F0FFF4', end_color='F0FFF4', fill_type='solid')
                            else:  # Question column
                                cell.fill = PatternFill(start_color='FAFAFA', end_color='FAFAFA', fill_type='solid')
            
            # Format summary sheet
            if 'Summary' in writer.sheets:
                summary_worksheet = writer.sheets['Summary']
                for column in summary_worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if cell.value:
                                cell_length = len(str(cell.value))
                                if cell_length > max_length:
                                    max_length = cell_length
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 30)
                    summary_worksheet.column_dimensions[column_letter].width = max(adjusted_width, 12)
        
        print(f"✅ Excel file created successfully: {output_file}")
        print(f"📊 Sheets created:")
        print(f"   - Question_Comparison: Side-by-side responses")
        print(f"   - Summary: Performance metrics comparison") 
        print(f"   - Metadata: Session information")
        
        return output_file
        
    except Exception as e:
        print(f"❌ Error creating Excel file: {e}")
        return None


def main():
    parser = argparse.ArgumentParser(description='Convert Multi-Prompt Evaluation Results to Excel')
    parser.add_argument('--input', required=True, help='Multi-prompt evaluation JSON results file')
    parser.add_argument('--output', help='Output Excel filename (auto-generated if not specified)')
    
    args = parser.parse_args()
    
    if not os.path.exists(args.input):
        print(f"❌ Input file not found: {args.input}")
        return 1
    
    try:
        output_file = convert_multi_prompt_to_excel(args.input, args.output)
        
        if output_file:
            print(f"🎉 Conversion completed successfully!")
            print(f"📄 Excel file: {output_file}")
            return 0
        else:
            return 1
            
    except Exception as e:
        print(f"❌ Conversion failed: {e}")
        return 1


if __name__ == "__main__":
    exit(main())